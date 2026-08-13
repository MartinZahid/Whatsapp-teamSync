# -*- coding: utf-8 -*-
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

SALIDA = "Cronograma-Gantt-WhatsApp-Team-Sync.xlsx"

FECHA_INICIO = date(2026, 8, 7)

COLORES = {
    "Análisis": "3B82F6",
    "Diseño y configuración": "8B5CF6",
    "Desarrollo": "22C55E",
    "Integración y pruebas": "F59E0B",
    "Documentación y cierre": "EF4444",
}

ACTIVIDADES = [
    ("Revisión del estado del arte y análisis de soluciones similares", "Análisis"),
    ("Levantamiento de requerimientos con el equipo de atención al cliente", "Análisis"),
    ("Diseño de la arquitectura de la solución y del protocolo", "Diseño y configuración"),
    ("Configuración del entorno de desarrollo (Vite, TypeScript, Node.js, Git)", "Diseño y configuración"),
    ("Desarrollo del servidor WebSocket (presencia, rooms, heartbeat)", "Desarrollo"),
    ("Desarrollo de la extensión de Chrome: popup y página de opciones", "Desarrollo"),
    ("Desarrollo del content script: panel flotante y detector de contacto", "Desarrollo"),
    ("Detección automática de chats y sincronización entre pestañas", "Desarrollo"),
    ("Desarrollo de la base de datos SQLite (sesiones y eventos)", "Desarrollo"),
    ("Desarrollo del portal web de monitoreo (dashboard)", "Desarrollo"),
    ("Integración del protocolo y pruebas de conexión y reconexión", "Integración y pruebas"),
    ("Pruebas unitarias y corrección de errores", "Integración y pruebas"),
    ("Pruebas funcionales con un equipo real de atención al cliente", "Integración y pruebas"),
    ("Pruebas de rendimiento y carga del servidor", "Integración y pruebas"),
    ("Capacitación a usuarios del sistema", "Documentación y cierre"),
    ("Elaboración de manuales de instalación y uso", "Documentación y cierre"),
    ("Despliegue del sistema y puesta en producción", "Documentación y cierre"),
    ("Documentación técnica y elaboración del reporte final", "Documentación y cierre"),
]

programa = []
d = FECHA_INICIO
for nombre, fase in ACTIVIDADES:
    fin = d + timedelta(days=6)
    programa.append((nombre, d, fin, fase))
    d = fin + timedelta(days=1)

FECHA_FIN = programa[-1][2]
N = len(programa)

wb = Workbook()
ws = wb.active
ws.title = "Cronograma"

azul = "1F3B73"
gris = "F2F2F2"
borde = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def fill(cell, color):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

ULT_COL = 6 + N
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ULT_COL)
ws.cell(1, 1, "CRONOGRAMA PRELIMINAR DE ACTIVIDADES - WhatsApp Team Sync").font = Font(bold=True, size=14, color="FFFFFF")
fill(ws.cell(1, 1), azul)

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ULT_COL)
ws.cell(2, 1, "Periodo: del %s al %s  |  %d semanas (una semana por actividad, consecutivas sin dejar dias vacios)" % (
    FECHA_INICIO.strftime("%d/%m/%Y"), FECHA_FIN.strftime("%d/%m/%Y"), N)).font = Font(italic=True, size=10)

HEADER = 4
encabezados = ["No.", "Actividad", "Semana", "Fecha inicial", "Fecha final", "Fase"]
for c, t in enumerate(encabezados, 1):
    celda = ws.cell(HEADER, c, t)
    celda.font = Font(bold=True, color="FFFFFF")
    fill(celda, azul)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for k in range(N):
    celda = ws.cell(HEADER, 7 + k, "%d\n%s" % (k + 1, programa[k][1].strftime("%d/%m")))
    celda.font = Font(bold=True, size=8, color="595959")
    fill(celda, gris)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = borde

for i, (nombre, ini, fin, fase) in enumerate(programa):
    fila = HEADER + 1 + i
    ws.cell(fila, 1, i + 1).alignment = Alignment(horizontal="center")
    ws.cell(fila, 2, nombre)
    ws.cell(fila, 3, "Semana %d" % (i + 1)).alignment = Alignment(horizontal="center")
    ws.cell(fila, 4, ini.strftime("%d/%m/%Y")).alignment = Alignment(horizontal="center")
    ws.cell(fila, 5, fin.strftime("%d/%m/%Y")).alignment = Alignment(horizontal="center")
    ws.cell(fila, 6, fase).alignment = Alignment(horizontal="center")
    for c in range(1, 7):
        ws.cell(fila, c).border = borde
    for k in range(N):
        celda = ws.cell(fila, 7 + k)
        celda.border = borde
        if k == i:
            fill(celda, COLORES[fase])
            celda.value = i + 1
            celda.font = Font(bold=True, color="FFFFFF", size=8)
            celda.alignment = Alignment(horizontal="center")

LEYENDA = 4 + N + 2
ws.cell(LEYENDA, 1, "Leyenda de fases:").font = Font(bold=True)
for j, (fase, color) in enumerate(COLORES.items()):
    fila = LEYENDA + 1 + j
    fill(ws.cell(fila, 1), color)
    ws.cell(fila, 2, fase).font = Font(size=10)

RESUMEN = LEYENDA + len(COLORES) + 2
ws.cell(RESUMEN, 1, "Total: %d semanas (4 meses)   |   Inicio: %s   |   Fin: %s" % (
    N, FECHA_INICIO.strftime("%d/%m/%Y"), FECHA_FIN.strftime("%d/%m/%Y"))).font = Font(bold=True, size=10)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 62
ws.column_dimensions["C"].width = 10
for col in ("D", "E"):
    ws.column_dimensions[col].width = 13
ws.column_dimensions["F"].width = 22
for k in range(N):
    ws.column_dimensions[get_column_letter(7 + k)].width = 7.5

ws.freeze_panes = "G5"
ws.row_dimensions[HEADER].height = 32

ws2 = wb.create_sheet("Gantt (grafica)")
ws2.cell(1, 2, "Actividad").font = Font(bold=True)
ws2.cell(1, 3, "Dias desde inicio").font = Font(bold=True)
ws2.cell(1, 4, "Duracion (dias)").font = Font(bold=True)
proyecto_inicio = FECHA_INICIO
for i, (nombre, ini, fin, fase) in enumerate(programa):
    fila = 2 + i
    ws2.cell(fila, 2, nombre)
    ws2.cell(fila, 3, (ini - proyecto_inicio).days)
    ws2.cell(fila, 4, (fin - ini).days + 1)

chart = BarChart()
chart.type = "bar"
chart.grouping = "stacked"
chart.overlap = 100
chart.gapWidth = 20
chart.title = "Diagrama de Gantt - Cronograma preliminar"
chart.style = 11

cats = Reference(ws2, min_col=2, min_row=2, max_row=1 + N)
data = Reference(ws2, min_col=3, max_col=4, min_row=1, max_row=1 + N)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

chart.series[0].graphicalProperties.noFill = True
chart.series[0].graphicalProperties.line.noFill = True
chart.series[1].graphicalProperties.solidFill = "22C55E"

chart.y_axis.title = "Actividades"
chart.x_axis.title = "Dias"
chart.legend = None
chart.width = 34
chart.height = 20

ws2.add_chart(chart, "F1")

wb.save(SALIDA)
print("OK ->", SALIDA)
print("Periodo:", FECHA_INICIO.strftime("%d/%m/%Y"), "al", FECHA_FIN.strftime("%d/%m/%Y"), "|", N, "semanas")
